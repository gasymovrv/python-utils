import os
from collections import defaultdict
from datetime import datetime

import matplotlib.pyplot as plt
import openpyxl

"""
This script reads headache data from Excel files and generates various statistics.

The script performs the following functions:
1. Reads data from multiple Excel files and aggregates it into a list of dictionaries.
2. Calculates the number of headaches per month.
3. Plots a number of medications used per month.
"""


def read_headache_data(file_paths):
    """
    Reads headache data from a list of Excel file paths and aggregates it.

    Args:
        file_paths (list of str): List of file paths to Excel files containing headache data.

    Returns:
        list of tuples: Each tuple contains the observation date, pain description, and a dictionary of medications with their usage.
    """
    aggregated_data = []

    for file_path in file_paths:
        workbook = openpyxl.load_workbook(file_path, data_only=True)

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            for row in sheet.iter_rows(min_row=2, values_only=True):
                if len(row) < 2:
                    continue

                # Validate the first column to ensure it's a date
                observation_date = row[0]
                if not isinstance(observation_date, datetime):
                    try:
                        observation_date = datetime.strptime(str(observation_date), "%Y-%m-%d")
                    except ValueError:
                        continue  # Skip row if the first column isn't a valid date

                # Ensure the second column (pain description) is not empty
                pain_description = row[1]
                if not pain_description:
                    continue  # Skip row if pain description is empty

                # Extract other columns (medications)
                medications = {}
                for i, value in enumerate(row[2:], start=2):  # Start from the third column
                    if value:  # Only include non-empty values
                        medications[sheet.cell(row=1, column=i + 1).value] = value

                # Append to the result list
                aggregated_data.append((observation_date, pain_description, medications))

    return aggregated_data


def plot_monthly_headache_trends(aggregated_data):
    """
    Plots the number of headaches per month based on the aggregated data.

    Args:
        aggregated_data (list of tuples): Aggregated headache data including observation date and pain description.
    """
    # Initialize a dictionary to count headaches per (year, month)
    month_counts = defaultdict(int)

    for observation_date, pain_description, _ in aggregated_data:
        year_month = (observation_date.year, observation_date.month)

        # Increment the count for the corresponding (year, month)
        month_counts[year_month] += 1

    # Prepare the data for plotting
    sorted_year_months = sorted(month_counts.keys())
    counts = [month_counts[year_month] for year_month in sorted_year_months]
    labels = [f"{year}-{month:02d}" for year, month in sorted_year_months]

    # Plot the data
    plt.figure(figsize=(12, 6))
    plt.plot(labels, counts, marker='o', linestyle='-', color='b')
    plt.title("Количество головных болей по месяцам")
    plt.xlabel("Время (Год-Месяц)")
    plt.ylabel("Количество")
    plt.xticks(rotation=45, ha='right')
    plt.grid(True)
    plt.tight_layout()
    plt.show()


def plot_medication_usage(aggregated_data):
    """
    Plots the usage trends of medications over time based on the aggregated data.

    Args:
        aggregated_data (list of tuples): Aggregated headache data including observation date and medications.
    """
    # Dictionary to store medication usage per (year-month)
    medication_usage = defaultdict(lambda: defaultdict(float))

    # Set to store unique Year-Month labels
    unique_time_labels = set()

    for observation_date, _, medications in aggregated_data:
        # Convert observation date to Year-Month format
        time_label = observation_date.strftime("%Y-%m")
        unique_time_labels.add(time_label)

        for med, value in medications.items():
            if med != "Результат" and isinstance(value, (int, float)):  # Only numeric values
                medication_usage[time_label][med] += value  # Sum medication usage per month

    # Sort time labels chronologically
    sorted_time_labels = sorted(unique_time_labels)

    # Get all unique medication names
    all_medications = set(med for data in medication_usage.values() for med in data)

    # Prepare data for plotting
    medication_trends = {med: [] for med in all_medications}
    medication_trends["Общее кол-во"] = []

    for date in sorted_time_labels:
        for med in all_medications:
            medication_trends[med].append(medication_usage[date].get(med, 0))  # Fill missing with 0
        medication_trends["Общее кол-во"].append(sum(medication_usage[date].values()))

    # Plot the medication usage trends
    plt.figure(figsize=(12, 6))

    for med, values in medication_trends.items():
        plt.plot(sorted_time_labels, values, marker="o", linestyle="-", label=med)

    plt.title("Количество принимаемых лекарств по месяцам")
    plt.xlabel("Время (Год-Месяц)")
    plt.ylabel("Количество")
    plt.xticks(rotation=45, ha="right")
    plt.legend(title="Название лекарства")
    plt.grid(True)
    plt.tight_layout()
    plt.show()


folder_path = "D:\\GoogleDisk\\Health\\Ruslan\\Мигрень"  # Folder containing yearly XLSX files
xlsx_files = [os.path.join(folder_path, f) for f in os.listdir(folder_path) if f.endswith(".xlsx")]

result = read_headache_data(xlsx_files)
# for r in result:
#     print(r)

plot_monthly_headache_trends(result)
plot_medication_usage(result)
